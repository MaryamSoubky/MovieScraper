from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException, WebDriverException
import time
import random
import openpyxl
import os
import json

CHROME_DRIVER_PATH = r'C:\Users\soubk\OneDrive\Desktop\chromedriver-win64\chromedriver.exe'
OUTPUT_FILE = 'IMDB_Top_250_Movies_With_Genres.xlsx'
CHECKPOINT_FILE = 'scrape_checkpoint.json'
MAX_RETRIES = 3
SLEEP_RANGE = (1.0, 2.5)


def setup_driver():
    chrome_options = Options()
    chrome_options.add_argument("--disable-blink-features=AutomationControlled")
    chrome_options.add_argument("--start-maximized")
    chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
    chrome_options.add_experimental_option('useAutomationExtension', False)
    chrome_options.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
    )
    service = Service(executable_path=CHROME_DRIVER_PATH)
    driver = webdriver.Chrome(service=service, options=chrome_options)
    # Mask webdriver flag
    driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
    return driver


def load_or_create_excel():

    if os.path.exists(OUTPUT_FILE):
        excel = openpyxl.load_workbook(OUTPUT_FILE)
        sheet = excel.active
        existing_rows = sheet.max_row - 1  # subtract header
        print(f"Resuming from existing file with {existing_rows} movies already saved.")
    else:
        excel = openpyxl.Workbook()
        sheet = excel.active
        sheet.title = 'IMDB Top 250 Movies'
        sheet.append(['Rank', 'Name', 'Year', 'Rating', 'Genre'])
        existing_rows = 0
    return excel, existing_rows


def load_checkpoint():
    if os.path.exists(CHECKPOINT_FILE):
        with open(CHECKPOINT_FILE, 'r') as f:
            return json.load(f)
    return {}


def save_checkpoint(data: dict):
    with open(CHECKPOINT_FILE, 'w') as f:
        json.dump(data, f)


def scroll_page(driver):
    """Scroll to bottom to trigger lazy-loaded content"""
    last_height = driver.execute_script("return document.body.scrollHeight")
    while True:
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(random.uniform(1.5, 3))
        new_height = driver.execute_script("return document.body.scrollHeight")
        if new_height == last_height:
            break
        last_height = new_height


def get_movie_links(soup):
    movie_links = []
    movies_list = soup.find(
        'ul',
        class_="ipc-metadata-list ipc-metadata-list--dividers-between sc-e22973a9-0 "
               "khSCXM compact-list-view ipc-metadata-list--base"
    )
    if movies_list:
        for movie in movies_list.find_all('li', class_="ipc-metadata-list-summary-item"):
            link = movie.find('a', class_="ipc-title-link-wrapper")
            if link:
                full_link = "https://www.imdb.com" + link['href'].split('?')[0]
                movie_links.append(full_link)
    return movie_links


def extract_movie_info(movie_item):
    title_element = movie_item.find('a', class_="ipc-title-link-wrapper")
    full_title = title_element.text.strip()  
    if '. ' in full_title:
        rank, name = full_title.split('. ', 1)
    else:
        rank = "N/A"
        name = full_title

    metadata = movie_item.find('div', class_="cli-title-metadata")
    year = metadata.find('span').text.strip() if metadata else "N/A"

    rating = movie_item.find('span', class_="ipc-rating-star--rating")
    rating_value = rating.text.strip() if rating else "N/A"

    return {'rank': rank, 'name': name, 'year': year, 'rating': rating_value}


def scrape_genres(driver, url, retries=MAX_RETRIES):
    for attempt in range(1, retries + 1):
        try:
            driver.get(url)

            try:
                WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located(
                        (By.CSS_SELECTOR, "[data-testid='genres'], .ipc-chip-list--base")
                    )
                )
            except TimeoutException:
                WebDriverWait(driver, 5).until(
                    EC.presence_of_element_located(
                        (By.XPATH, "//*[contains(text(),'Genre') or contains(text(),'genres')]")
                    )
                )

            selectors = [
                "[data-testid='genres'] a",
                ".ipc-chip-list--base a",
                ".see-more.inline a[href*='genre']",
                "a[href*='genres=']"
            ]
            for selector in selectors:
                elements = driver.find_elements(By.CSS_SELECTOR, selector)
                if elements:
                    genres = ', '.join([el.text for el in elements if el.text])
                    if genres:
                        return genres

            return "N/A"

        except (TimeoutException, WebDriverException) as e:
            print(f"  Attempt {attempt}/{retries} failed for {url}: {str(e)[:100]}")
            if attempt < retries:
                time.sleep(random.uniform(3, 6))
        finally:
            time.sleep(random.uniform(*SLEEP_RANGE))

    return "N/A"


def main():
    excel, already_done = load_or_create_excel()
    sheet = excel.active
    checkpoint = load_checkpoint()
    driver = setup_driver()

    try:
        print("Loading IMDb Top 250 page...")
        driver.get('https://www.imdb.com/chart/top/')
        WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.CLASS_NAME, "ipc-metadata-list"))
        )

        print("Scrolling to load all movies...")
        scroll_page(driver)

        soup = BeautifulSoup(driver.page_source, 'html.parser')
        movie_links = get_movie_links(soup)
        movies_list = soup.find_all('li', class_="ipc-metadata-list-summary-item")

        total = min(len(movie_links), len(movies_list), 250)
        print(f"Found {total} movies. Starting genre scraping...")

        for i in range(already_done, total):
            link = movie_links[i]
            try:
                movie_info = extract_movie_info(movies_list[i])

                if str(i) in checkpoint:
                    genres = checkpoint[str(i)]
                    print(f"  [cached] {i+1}/{total}: {movie_info['name']}")
                else:
                    genres = scrape_genres(driver, link)
                    checkpoint[str(i)] = genres
                    save_checkpoint(checkpoint) 

                sheet.append([
                    movie_info['rank'],
                    movie_info['name'],
                    movie_info['year'],
                    movie_info['rating'],
                    genres
                ])
                
                if (i + 1) % 10 == 0:
                    excel.save(OUTPUT_FILE)
                    print(f"  💾 Auto-saved at movie {i+1}")

                print(f"Processed {i+1}/{total}: {movie_info['rank']}. {movie_info['name']} | {genres}")

            except Exception as e:
                print(f"Failed on movie {i+1}: {str(e)[:200]}")
                continue

    except Exception as e:
        print(f"Fatal error: {str(e)}")
    finally:
        driver.quit()
        excel.save(OUTPUT_FILE)
        excel.close()
    
        if os.path.exists(CHECKPOINT_FILE):
            os.remove(CHECKPOINT_FILE)
        print(f"Done. Data saved to {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
